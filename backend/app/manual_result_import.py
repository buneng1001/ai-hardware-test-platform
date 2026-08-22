import csv
import io
from datetime import UTC, datetime
from pathlib import Path
from zipfile import BadZipFile

from fastapi import APIRouter, HTTPException, Request, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.database import open_database
from app.manual_check_results import ManualCheckResultCommand, result_from_row
from app.run_models import ManualCheckResult

router = APIRouter(prefix="/api/runs/{run_id}/manual-check-results", tags=["manual check results"])
MAX_IMPORT_BYTES = 2 * 1024 * 1024
IMPORT_COLUMNS = ["name", "status", "actual_result", "notes", "executed_at"]
MANUAL_STATUSES = {"passed", "failed", "blocked", "not_run"}


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise HTTPException(status_code=422, detail="CSV 必须使用 UTF-8 编码") from error
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames != IMPORT_COLUMNS:
        raise HTTPException(status_code=422, detail=f"导入列必须是 {','.join(IMPORT_COLUMNS)}")
    return [{key: value or "" for key, value in row.items() if key is not None} for row in reader]


def _parse_excel(content: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise HTTPException(status_code=422, detail="Excel 文件无法读取") from error
    try:
        values = workbook.active.iter_rows(values_only=True)
        header = [str(value) if value is not None else "" for value in next(values, ())]
        if header != IMPORT_COLUMNS:
            raise HTTPException(status_code=422, detail=f"导入列必须是 {','.join(IMPORT_COLUMNS)}")
        return [
            {
                column: value.isoformat() if isinstance(value, datetime) else str(value) if value is not None else ""
                for column, value in zip(IMPORT_COLUMNS, row, strict=False)
            }
            for row in values
        ]
    finally:
        workbook.close()


def _validate_rows(rows: list[dict[str, str]]) -> list[ManualCheckResultCommand]:
    commands: list[ManualCheckResultCommand] = []
    errors: list[dict[str, str | int]] = []
    for row_number, row in enumerate(rows, start=2):
        row_errors: list[dict[str, str | int]] = []
        if not row["name"].strip():
            row_errors.append({"row": row_number, "field": "name", "message": "人工检查项名称不能为空"})
        elif len(row["name"]) > 120:
            row_errors.append({"row": row_number, "field": "name", "message": "人工检查项名称不能超过 120 字符"})
        if row["status"] not in MANUAL_STATUSES:
            row_errors.append(
                {
                    "row": row_number,
                    "field": "status",
                    "message": "状态必须是 passed、failed、blocked 或 not_run",
                }
            )
        if len(row["actual_result"]) > 2000:
            row_errors.append({"row": row_number, "field": "actual_result", "message": "实际结果不能超过 2000 字符"})
        if len(row["notes"]) > 2000:
            row_errors.append({"row": row_number, "field": "notes", "message": "备注不能超过 2000 字符"})
        if row["executed_at"]:
            try:
                datetime.fromisoformat(row["executed_at"])
            except ValueError:
                row_errors.append({"row": row_number, "field": "executed_at", "message": "执行时间格式无效"})
        if row_errors:
            errors.extend(row_errors)
            continue
        commands.append(
            ManualCheckResultCommand(
                name=row["name"],
                status=row["status"],
                actual_result=row["actual_result"] or None,
                notes=row["notes"] or None,
                executed_at=row["executed_at"] or None,
            )
        )
    if errors:
        raise HTTPException(status_code=422, detail=errors)
    return commands


def _insert_results(run_id: int, commands: list[ManualCheckResultCommand]) -> list[ManualCheckResult]:
    now = datetime.now(UTC).isoformat()
    result_ids: list[int] = []
    with open_database() as connection:
        if connection.execute("SELECT id FROM runs WHERE id = ?", (run_id,)).fetchone() is None:
            raise HTTPException(status_code=404, detail="运行记录不存在")
        for command in commands:
            cursor = connection.execute(
                """
                INSERT INTO manual_check_results (
                    run_id, name, status, actual_result, notes, executed_at, attachment, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, NULL, ?, ?)
                """,
                (
                    run_id,
                    command.name,
                    command.status,
                    command.actual_result,
                    command.notes,
                    command.executed_at.isoformat() if command.executed_at else None,
                    now,
                    now,
                ),
            )
            if cursor.lastrowid is not None:
                result_ids.append(cursor.lastrowid)
        placeholders = ",".join("?" for _ in result_ids)
        rows = connection.execute(
            f"SELECT * FROM manual_check_results WHERE id IN ({placeholders}) ORDER BY id",
            result_ids,
        ).fetchall()
    return [result_from_row(row) for row in rows]


@router.post("/import", response_model=list[ManualCheckResult], status_code=status.HTTP_201_CREATED)
async def import_manual_results(run_id: int, filename: str, request: Request) -> list[ManualCheckResult]:
    content = await request.body()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="导入文件不能超过 2 MiB")
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(content)
    elif suffix == ".xlsx":
        rows = _parse_excel(content)
    else:
        raise HTTPException(status_code=422, detail="导入文件必须是 CSV 或 XLSX")
    return _insert_results(run_id, _validate_rows(rows))
